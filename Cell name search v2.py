import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import numbers
wb = openpyxl.load_workbook('Summary.xlsx', read_only=False, keep_vba=False, data_only=True, keep_links=False)
print(wb.sheetnames)
sheetn = input("Enter the sheet number:" )
#sheet 2 mode
ws = wb.create_sheet()
ws.title = "Sheet " + sheetn + " Analysis"
ws.cell(row=1,column=1).value = wb.sheetnames[int(sheetn)-1]
ws.cell(row=2,column=1).value = 'Analysis'
ws.cell(row=4,column=1).value = 'Income (without DBCS and Donations)'
ws.cell(row=5,column=1).value = 'Variable costs of Paid surgeries'
ws.cell(row=6,column=1).value = 'Gross Margin (A-B)'
ws.cell(row=8,column=1).value = 'Fixed Cost'
ws.cell(row=9,column=1).value = 'Balance Fixed Cost to be covered (D-C)'
ws.cell(row=11,column=1).value = 'Variable Cost of Free surgeries'
ws.cell(row=13,column=1).value = 'Operating Deficit (E+F)'
ws.cell(row=15,column=1).value = 'DBCS'
ws.cell(row=16,column=1).value = 'Committed Donations (including SEF)'
ws.cell(row=18,column=1).value = 'Operating Deficit to be funded/ (Surplus) (G-H-I)'
ws.cell(row=20,column=1).value = 'Capital Replacement Fund- CRF'
ws.cell(row=22,column=1).value = 'Total Deficit Funding Requirement incl CRF/ (Surplus) (J+K)'
ws.cell(row=24,column=1).value = 'Capital expansion Spends'
ws.cell(row=25,column=1).value = 'Committed donations for Capital'
ws.cell(row=26,column=1).value = 'Need of funds/ (Surplus of funds) (M+N-O)'
ws.cell(row=29,column=1).value = 'Deficit shown'
ws.cell(row=31,column=1).value = 'Receivables at Year-end'
ws.cell(row=32,column=1).value = 'Payables at Year-end'
ws.cell(row=33,column=1).value = 'Bank Balance at Year-end'
ws.cell(row=35,column=1).value = 'Paid surgeries(nos)'
ws.cell(row=36,column=1).value = 'Free surgeries(nos)'
ws.cell(row=37,column=1).value = 'Revenue per Paid surgery '
ws.cell(row=39,column=1).value = 'Indicated Self sufficieny (thru internal operations)'
ws = wb[wb.sheetnames[int(sheetn)-1]]
#sheet 1 mode
tempr = int(input("Enter the (last + 1) cell row: "))
tempc = int(input("Enter the (last + 1) cell column: "))
#search for titles in sheet 1
for i in range(1,tempr):
    if ws.cell(row=i,column=1).value == 'OP income':
        opincome = i
    elif ws.cell(row=i,column=1).value == 'IP income':
            ipincome = i
    elif ws.cell(row=i,column=1).value =='Pharmacy Income (net of expenses)':
        pharmaincome = i
    elif ws.cell(row=i,column=1).value == 'Optical contribution (net of expenses)':
        opticcontrib = i
    elif ws.cell(row=i,column=1).value == 'DBCS Income':
        dbcsincome = i
    elif ws.cell(row=i,column=1).value == 'Training & other income':
        trainincome = i
    elif ws.cell(row=i,column=1).value == 'Opex Donations (including SEF USA)':
        opexdono = i
    elif ws.cell(row=i,column=1).value == 'Non Paying Variable cost':
        nopayvarcost = i
    elif ws.cell(row=i,column=1).value == 'Camp cost (Excluding HR, payroll/salaries, fixed costs)':
        campcost = i
    elif ws.cell(row=i,column=1).value == 'Paying Variable cost':
        payvarcost = i
    elif ws.cell(row=i,column=1).value == 'Fixed Cost (incl. Interest but excluding Depreciation)':
        fixedcost = i
    elif ws.cell(row=i,column=1).value == 'Outpatient Footfall (Paid; not incl. Review cases)':
        outpatfall = i
    elif ws.cell(row=i,column=1).value == 'Paid Surgeries':
        paidsurg = i
    elif ws.cell(row=i,column=1).value == 'Free Surgeries':
        freesurg = i
    elif ws.cell(row=i,column=1).value == '# of Main Camps':
        maincamps = i
    elif ws.cell(row=i,column=1).value == 'Total income':
        tincome = i
    elif ws.cell(row=i,column=1).value == 'Total Expenses':
        texpense = i
    elif ws.cell(row=i,column=1).value == 'Capital Spend so far in the Year':
        capspent = i
    elif ws.cell(row=i,column=1).value == 'Capital donations received so far':
        caprec = i
    elif ws.cell(row=i,column=1).value == 'Accounts receivables as at Quarter-end':
        accrec = i
    elif ws.cell(row=i,column=1).value == 'Accounts Payables as at Quarter-end':
        accpay = i
    elif ws.cell(row=i,column=1).value == 'Bank Balance at Quarter-end':
        bbqend = i
    elif ws.cell(row=i,column=1).value == 'Capital Replacement fund':
        crfu = i


    
#create extra titles in sheet 1
ws.cell(row=(texpense+1),column=1).value = 'Surplus/Deficit'
ws.cell(row=tempr,column=1).value = 'Self Sufficiency%'
ws.cell(row=tempr+1,column=1).value = 'Outpatient Conversion Rate'
ws.cell(row=tempr+2,column=1).value = 'Cost/Camp'
ws.cell(row=tempr+3,column=1).value = 'Var Cost/Paid'
ws.cell(row=tempr+4,column=1).value = 'Var Cost/Free'
ws.cell(row=tempr+5,column=1).value = 'Fixed Cost/Month'
ws.cell(row=tempr+6,column=1).value = 'Revenue/Paid'
revpay  = tempr+6

#loop for both sheets 

for j in range(2,tempc):
    ws = wb[wb.sheetnames[int(sheetn)-1]]
    ws.cell(row=tincome,column=j).value = ws.cell(row=opincome,column=j).value + ws.cell(row=ipincome,column=j).value + ws.cell(row=pharmaincome,column=j).value + ws.cell(row=opticcontrib,column=j).value + ws.cell(row=dbcsincome,column=j).value + ws.cell(row=trainincome,column=j).value + ws.cell(row=opexdono,column=j).value 
    ws.cell(row=texpense,column=j).value = ws.cell(row=nopayvarcost,column=j).value+ws.cell(row=campcost,column=j).value +ws.cell(row=payvarcost,column=j).value+ ws.cell(row=fixedcost,column=j).value
    ws.cell(row=(texpense+1),column=j).value = (ws.cell(row=tincome,column=j).value) - (ws.cell(row=texpense,column=j).value)
    ws.cell(row=(texpense+1),column=j).font = Font(bold=True)
    if ws.cell(row=texpense,column=j).value==0 :
        ws.cell(row=tempr,column=j).value = 0
    else:
        ws.cell(row=tempr,column=j).value = (ws.cell(row=tincome,column=j).value  - ws.cell(row=opexdono,column=j).value - ws.cell(row=dbcsincome,column=j).value ) / (ws.cell(row=texpense,column=j).value)
    ws.cell(row=tempr,column=j).value = str(int((ws.cell(row=tempr,column=j).value)*100)) + '%'
    ws.cell(row=tempr,column=j).font = Font(bold=True)
    if ws.cell(row=outpatfall,column=j).value==0:
        ws.cell(row=tempr+1,column=j).value = 0
    else:
        ws.cell(row=tempr+1,column=j).value = str(int(int(ws.cell(row=paidsurg,column=j).value)* 100 / ws.cell(row=outpatfall,column=j).value))+'%'
        ws.cell(row=tempr+1,column=j).font = Font(bold=True)
    if ws.cell(row=maincamps,column=j).value==0:
        ws.cell(row=tempr+2,column=j).value = 0
    else:
        ws.cell(row=tempr+2,column=j).value = int(ws.cell(row=campcost,column=j).value/ws.cell(row=maincamps,column=j).value)
        ws.cell(row=tempr+2,column=j).font = Font(bold=True)
    if ws.cell(row=payvarcost,column=j).value==0:
        ws.cell(row=tempr+3,column=j).value = 0
    else:
        ws.cell(row=tempr+3,column=j).value = int(ws.cell(row=payvarcost,column=j).value/ws.cell(row=paidsurg,column=j).value)
        ws.cell(row=tempr+3,column=j).font = Font(bold=True)
    if ws.cell(row=freesurg,column=j).value==0:
        ws.cell(row=tempr+4,column=j).value = 0
    else:
        ws.cell(row=tempr+4,column=j).value = int((ws.cell(row=nopayvarcost,column=j).value + ws.cell(row=campcost,column=j).value)/ws.cell(row=freesurg,column=j).value)
        ws.cell(row=tempr+4,column=j).font = Font(bold=True)
    ws.cell(row=tempr+5,column=j).value = int(ws.cell(row=fixedcost,column=j).value/3)
    ws.cell(row=tempr+5,column=j).font = Font(bold=True)
    if ws.cell(row=paidsurg,column=j).value==0:
        ws.cell(row=tempr+6,column=j).value = 0
    else:
        ws.cell(row=tempr+6,column=j).value = int(ws.cell(row=ipincome,column=j).value/ws.cell(row=paidsurg,column=j).value)
        ws.cell(row=tempr+6,column=j).font = Font(bold=True)

    #save data from sheet 1 
    totalincome = (ws.cell(row=tincome,column=j).value) - (ws.cell(row=opexdono,column=j).value) - (ws.cell(row=dbcsincome,column=j).value)
    varpaidsurg = (ws.cell(row=payvarcost,column=j).value)
    fixcos = (ws.cell(row=fixedcost,column=j).value)
    varfreesurg = (ws.cell(row=campcost,column=j).value) + (ws.cell(row=nopayvarcost,column=j).value)
    dbcs2 = (ws.cell(row=dbcsincome,column=j).value)
    commdono = (ws.cell(row=opexdono,column=j).value)
    capitalspent = int((ws.cell(row=capspent,column=j).value) or 0) 
    capitalrecie = int((ws.cell(row=caprec,column=j).value) or 0)
    recatend = (ws.cell(row=accrec,column=j).value)
    payatend = (ws.cell(row=accpay,column=j).value)
    bankbal = (ws.cell(row=bbqend,column=j).value)
    psurg = (ws.cell(row=paidsurg,column=j).value)
    fsurg = (ws.cell(row=freesurg,column=j).value)
    rpsurg = (ws.cell(row=revpay,column=j).value)
    crfund = (ws.cell(row=crfu,column=j).value)


    #sheet 2 mode
    ws = wb["Sheet " + sheetn + " Analysis"]
    (ws.cell(row=4,column=j).value) = totalincome
    (ws.cell(row=5,column=j).value) = varpaidsurg
    (ws.cell(row=6,column=j).value) = totalincome - varpaidsurg
    (ws.cell(row=8,column=j).value) = fixcos
    (ws.cell(row=9,column=j).value) = fixcos - totalincome - varpaidsurg 
    (ws.cell(row=11,column=j).value) = varfreesurg
    (ws.cell(row=13,column=j).value) = varfreesurg + fixcos - totalincome - varpaidsurg
    (ws.cell(row=15,column=j).value) = dbcs2
    (ws.cell(row=16,column=j).value) = commdono
    (ws.cell(row=18,column=j).value) = varfreesurg + fixcos - totalincome - varpaidsurg - dbcs2 - commdono
    (ws.cell(row=20,column=j).value) = crfund
    (ws.cell(row=22,column=j).value) = (ws.cell(row=18,column=j).value)
    (ws.cell(row=24,column=j).value) = capitalspent
    (ws.cell(row=25,column=j).value) = capitalrecie
    (ws.cell(row=26,column=j).value) = (ws.cell(row=22,column=j).value) + capitalspent - capitalrecie
    (ws.cell(row=29,column=j).value) = '-'
    (ws.cell(row=31,column=j).value) = recatend
    (ws.cell(row=32,column=j).value) = payatend
    (ws.cell(row=33,column=j).value) = int(bankbal)
    (ws.cell(row=35,column=j).value) = psurg
    (ws.cell(row=36,column=j).value) = fsurg
    (ws.cell(row=37,column=j).value) = rpsurg
    (ws.cell(row=39,column=j).value) = str(int(totalincome * 100  / (varpaidsurg + fixcos + varfreesurg))) + '%'

wb.save('newfile.xlsx')
