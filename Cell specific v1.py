import openpyxl
from openpyxl.styles import Font
wb = openpyxl.load_workbook('Summary.xlsx', read_only=False, keep_vba=False, data_only=True, keep_links=False)
print(wb.sheetnames)
tempn = input("Enter the sheet number: ")
ws = wb[wb.sheetnames[int(tempn)-1]]


for j in range(2,6):
    total = 0
    for i in range(4,11):
        data = ws.cell(row=i, column=j).value
        total += data;
    ws.cell(row=11,column=j).value = total


for j in range(2,6):
    total = 0
    for i in range(13,17):
        data = ws.cell(row=i,column=j).value
        total += data;
    ws.cell(row=17,column=j).value = total


ws.cell(row=18,column=1).value = 'Surplus/Deficit'
for j in range(2,6):
    ws.cell(row=18,column=j).value = (ws.cell(row=11,column=j).value) - (ws.cell(row=17,column=j).value)
    ws.cell(row=18,column=j).font = Font(bold=True)


ws.cell(row=22,column=1).value = 'Self Suffiency%'
for j in range(2,6):
    if ws.cell(row=17,column=j).value==0:
        ws.cell(row=22,column=j).value = 0
    else:
        ws.cell(row=22,column=j).value = (ws.cell(row=11,column=j).value - ws.cell(row=10,column=j).value - ws.cell(row=8,column=j).value)/ws.cell(row=17,column=j).value
        ws.cell(row=22,column=j).value = str(int((ws.cell(row=22,column=j).value)*100)) + '%'
    ws.cell(row=22,column=j).font = Font(bold=True)


ws.cell(row=24,column=1).value = 'Outpatient Conversion Rate'
for j in range(2,6):
    if ws.cell(row=25,column=j).value==0:
        ws.cell(row=24,column=j).value = 0
    else:
        ws.cell(row=24,column=j).value = str(int((ws.cell(row=26,column=j).value * 100) / (ws.cell(row=25,column=j).value)))+'%'
        ws.cell(row=24,column=j).font = Font(bold=True)


ws.cell(row=36,column=1).value = 'Cost/Camp'
for j in range(2,6):
    if ws.cell(row=28,column=j).value==0:
        ws.cell(row=36,column=j).value = 0
    else:
        ws.cell(row=36,column=j).value = int(ws.cell(row=14,column=j).value/ws.cell(row=28,column=j).value)
        ws.cell(row=36,column=j).font = Font(bold=True)


ws.cell(row=45,column=1).value = 'Var Cost/Paid'
for j in range(2,6):
    if ws.cell(row=15,column=j).value==0:
        ws.cell(row=45,column=j).value = 0
    else:
        ws.cell(row=45,column=j).value = int(ws.cell(row=15,column=j).value/ws.cell(row=26,column=j).value)
        ws.cell(row=45,column=j).font = Font(bold=True)

ws.cell(row=46,column=1).value = 'Var Cost/Free'
for j in range(2,6):
    if ws.cell(row=27,column=j).value==0:
        ws.cell(row=46,column=j).value = 0
    else:
        ws.cell(row=46,column=j).value = int(ws.cell(row=13,column=j).value + ws.cell(row=14,column=j).value/ws.cell(row=27,column=j).value)
        ws.cell(row=46,column=j).font = Font(bold=True)


ws.cell(row=47,column=1).value = 'Fixed Cost/Month'
for j in range(2,6):
        ws.cell(row=47,column=j).value = int(ws.cell(row=16,column=j).value/12)
        ws.cell(row=47,column=j).font = Font(bold=True)


ws.cell(row=48,column=1).value = 'Revenue/Paid'
for j in range(2,6):
    if ws.cell(row=26,column=j).value==0:
        ws.cell(row=48,column=j).value = 0
    else:
        ws.cell(row=48,column=j).value = int(ws.cell(row=5,column=j).value/ws.cell(row=26,column=j).value)
        ws.cell(row=48,column=j).font = Font(bold=True)


ws.cell(row=49,column=1).value = 'Micro-Conversion Rate(Free)'
for j in range(2,6):
    if ws.cell(row=30,column=j).value==0:
        ws.cell(row=49,column=j).value = 0
    else:
        ws.cell(row=49,column=j).value = int(ws.cell(row=32,column=j).value/ws.cell(row=30,column=j).value)
        ws.cell(row=49,column=j).font = Font(bold=True)


ws.cell(row=50,column=1).value = 'Micro-Conversion Rate(Paid)'
for j in range(2,6):
    if ws.cell(row=34,column=j).value==int():
        if ws.cell(row=34,column=j).value==0:
            ws.cell(row=50,column=j).value = 0
        else:
            ws.cell(row=50,column=j).value = int(ws.cell(row=35,column=j).value/ws.cell(row=34,column=j).value)
            ws.cell(row=50,column=j).font = Font(bold=True)

wb.save('newfile.xlsx')
